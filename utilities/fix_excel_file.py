# fix_excel_file.py
"""
Fixes problematic Excel files that can't be read by openpyxl
Resaves them in a clean format
Project Path: C:\\users\\rlzim\\PycharmProjects\\ESE_Policy4900
"""

import pandas as pd
from pathlib import Path
from colorama import init, Fore, Style
import sys
import warnings
import shutil

init(autoreset=True)
warnings.filterwarnings('ignore')


def fix_excel_file(input_path):
    """
    Read and resave an Excel file to fix formatting issues
    """
    print(f"\n{Fore.CYAN}Attempting to fix: {input_path}{Style.RESET_ALL}")

    try:
        # Create backup first
        backup_path = input_path.parent / f"{input_path.stem}_backup{input_path.suffix}"
        if not backup_path.exists():
            shutil.copy2(input_path, backup_path)
            print(f"{Fore.YELLOW}Backup created: {backup_path.name}{Style.RESET_ALL}")

        df = None
        error_messages = []

        # Method 1: Try with openpyxl engine (default for .xlsx)
        try:
            print(f"Method 1: Reading with openpyxl engine...")
            df = pd.read_excel(input_path, engine='openpyxl')
            print(f"{Fore.GREEN}✓ Success with openpyxl{Style.RESET_ALL}")
        except Exception as e:
            error_messages.append(f"openpyxl: {str(e)[:100]}")

        # Method 2: Try without specifying engine (let pandas decide)
        if df is None:
            try:
                print(f"Method 2: Reading with pandas auto-detect...")
                df = pd.read_excel(input_path)
                print(f"{Fore.GREEN}✓ Success with auto-detect{Style.RESET_ALL}")
            except Exception as e:
                error_messages.append(f"auto: {str(e)[:100]}")

        # Method 3: Try reading as CSV (sometimes Excel files are actually CSV)
        if df is None:
            try:
                print(f"Method 3: Trying as CSV...")
                df = pd.read_csv(input_path)
                print(f"{Fore.GREEN}✓ Read as CSV{Style.RESET_ALL}")
            except Exception as e:
                error_messages.append(f"csv: {str(e)[:100]}")

        # Method 4: Use openpyxl directly with data_only
        if df is None:
            try:
                print(f"Method 4: Using openpyxl directly (values only)...")
                from openpyxl import load_workbook
                wb = load_workbook(input_path, data_only=True, read_only=True)
                ws = wb.active

                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)

                df = pd.DataFrame(data[1:], columns=data[0])
                wb.close()
                print(f"{Fore.GREEN}✓ Success with direct openpyxl{Style.RESET_ALL}")
            except Exception as e:
                error_messages.append(f"direct openpyxl: {str(e)[:100]}")

        # Method 5: Force read with minimal options
        if df is None:
            try:
                print(f"Method 5: Minimal read attempt...")
                import openpyxl
                # Temporarily modify openpyxl's parser to be more lenient
                original_warning = openpyxl.xml.functions.LXML
                openpyxl.xml.functions.LXML = False

                df = pd.read_excel(
                    input_path,
                    engine='openpyxl',
                    dtype=str  # Read everything as string to avoid type issues
                )

                openpyxl.xml.functions.LXML = original_warning
                print(f"{Fore.GREEN}✓ Success with minimal options{Style.RESET_ALL}")
            except Exception as e:
                error_messages.append(f"minimal: {str(e)[:100]}")

        if df is None:
            print(f"\n{Fore.RED}Could not read file with any method{Style.RESET_ALL}")
            print(f"\n{Fore.YELLOW}Errors encountered:{Style.RESET_ALL}")
            for msg in error_messages:
                print(f"  - {msg}")

            # Try manual recovery
            print(f"\n{Fore.YELLOW}Attempting manual recovery...{Style.RESET_ALL}")
            return try_manual_recovery(input_path, backup_path)

        # Show what we found
        print(f"\n{Fore.GREEN}File contents successfully read:{Style.RESET_ALL}")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")
        print(f"  Headers: {list(df.columns[:5])}...")

        # Clean the data
        print(f"\n{Fore.CYAN}Cleaning data...{Style.RESET_ALL}")

        # Remove any completely empty rows
        df = df.dropna(how='all')

        # Ensure proper column names
        expected_cols = [
            "School of Instruction", "FISH Number", "Room", "FISH List",
            "Total Student Count", "# of Students Opt In", "# of Students Opt Out",
            "# of Students No Response", "% Opt In", "% Opt Out", "% No Response",
            "# of ESE Students"
        ]

        if len(df.columns) >= len(expected_cols):
            df.columns = expected_cols + list(df.columns[len(expected_cols):])

        # Convert numeric columns
        numeric_cols = ["Total Student Count", "# of Students Opt In",
                        "# of Students Opt Out", "# of Students No Response",
                        "# of ESE Students"]

        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

        # Convert percentage columns
        pct_cols = ["% Opt In", "% Opt Out", "% No Response"]
        for col in pct_cols:
            if col in df.columns:
                # Remove % sign if present and convert
                df[col] = df[col].astype(str).str.rstrip('%')
                df[col] = pd.to_numeric(df[col], errors='coerce')
                # If values are > 1, they're percentages, divide by 100
                mask = df[col] > 1
                df.loc[mask, col] = df.loc[mask, col] / 100

        # Save clean version
        output_path = input_path
        print(f"\n{Fore.CYAN}Saving clean version...{Style.RESET_ALL}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Policy 4900 - Classroom Percent', index=False)

        print(f"{Fore.GREEN}✓ Fixed file saved as: {output_path.name}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}✓ Original backed up as: {backup_path.name}{Style.RESET_ALL}")

        return True

    except Exception as e:
        print(f"{Fore.RED}Critical error: {e}{Style.RESET_ALL}")
        return False


def try_manual_recovery(input_path, backup_path):
    """Last resort: try to recover data manually"""
    print(f"\n{Fore.YELLOW}Manual recovery attempt...{Style.RESET_ALL}")

    try:
        # Try using xlrd for older Excel formats
        try:
            import xlrd
            print("Trying with xlrd library...")
            book = xlrd.open_workbook(input_path)
            sheet = book.sheet_by_index(0)

            data = []
            for row in range(sheet.nrows):
                data.append(sheet.row_values(row))

            df = pd.DataFrame(data[1:], columns=data[0])

            # Save the recovered data
            with pd.ExcelWriter(input_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Policy 4900 - Classroom Percent', index=False)

            print(f"{Fore.GREEN}✓ Manual recovery successful!{Style.RESET_ALL}")
            return True

        except ImportError:
            print(f"{Fore.YELLOW}xlrd not installed. Install with: pip install xlrd{Style.RESET_ALL}")
        except Exception as e:
            print(f"xlrd failed: {e}")

        # If all else fails, provide instructions
        print(f"\n{Fore.YELLOW}Automatic recovery failed. Manual steps:{Style.RESET_ALL}")
        print(f"1. Open {input_path.name} in Excel")
        print(f"2. Click File → Save As")
        print(f"3. Choose 'Excel Workbook (*.xlsx)' as the format")
        print(f"4. Save with the same name (overwrite)")
        print(f"5. Run this script again")

        return False

    except Exception as e:
        print(f"{Fore.RED}Manual recovery failed: {e}{Style.RESET_ALL}")
        return False


def main():
    """Main function to fix problematic Excel files"""
    print(f"{Fore.CYAN}{'=' * 60}")
    print(f"{'Excel File Fixer for Policy 4900':^60}")
    print(f"{'=' * 60}{Style.RESET_ALL}")

    # Default path for new reports
    reports_dir = Path(r"C:\BCPS\ESE\Policy4900_Reports_New")

    if len(sys.argv) > 1:
        # File specified as argument
        file_path = Path(sys.argv[1])
        if file_path.exists():
            success = fix_excel_file(file_path)
            if not success:
                print(f"\n{Fore.YELLOW}Try installing xlrd:{Style.RESET_ALL}")
                print(f"pip install xlrd")
        else:
            print(f"{Fore.RED}File not found: {file_path}{Style.RESET_ALL}")
    else:
        # Look for problematic files in the reports directory
        print(f"\nLooking for Excel files in: {reports_dir}")

        excel_files = list(reports_dir.glob("*_Policy4900.xlsx"))
        unprocessed = [f for f in excel_files if
                       not f.name.endswith('.processed') and not f.name.endswith('_backup.xlsx')]

        if not unprocessed:
            print(f"{Fore.YELLOW}No unprocessed Policy4900 files found{Style.RESET_ALL}")
            return

        print(f"\nFound {len(unprocessed)} file(s):")
        for i, file in enumerate(unprocessed, 1):
            print(f"  {i}. {file.name}")

        if len(unprocessed) == 1:
            choice = 1
        else:
            choice = input(f"\nWhich file to fix? (1-{len(unprocessed)}): ")
            try:
                choice = int(choice)
            except:
                print(f"{Fore.RED}Invalid choice{Style.RESET_ALL}")
                return

        if 1 <= choice <= len(unprocessed):
            success = fix_excel_file(unprocessed[choice - 1])
            if not success:
                print(f"\n{Fore.YELLOW}Additional options:{Style.RESET_ALL}")
                print(f"1. Install xlrd: pip install xlrd")
                print(f"2. Open the file in Excel and Save As 'Excel Workbook'")
        else:
            print(f"{Fore.RED}Invalid choice{Style.RESET_ALL}")


if __name__ == "__main__":
    main()
    input(f"\n{Fore.CYAN}Press Enter to exit...{Style.RESET_ALL}")