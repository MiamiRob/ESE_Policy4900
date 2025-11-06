# task2_merge_to_master_enhanced.py
"""
TASK 2: Merge processed reports into the master tracking workbook
ENHANCED WITH RETENTION SYSTEM for installed/active cameras

Key Features:
- AUTOMATIC BACKUP of master file before each run (timestamped)
- Permanent retention of any camera with Active/Activated status or Installed/Existing status
- Retention file (Policy4900_Retained.csv) as backup persistence layer
- Automatic rehydration when loading master to prevent data loss
- Never drops installed/active rows even if missing from district CSV reports
- Missing cameras get properly classified labels (Missing - Camera Active, etc.)
- Detailed console output showing which schools/rooms changed with report dates

- Reads *_Policy4900_PROCESSED.csv files from paths.new_reports_dir
- Merges into Policy4900_Tracking_Master.xlsx in paths.master_dir/master_file
- Calculates Approval Status in Python from raw data
- Applies conditional formatting
- Honors manual Void overrides from paths.void_overrides_file

Manual Void rules:
- Overrides file headers: School of Instruction, Room, Mark if Void, Approval Status
- If a (School, Room) appears with either "Mark if Void" == "Void" OR "Approval Status" == "Void",
  the row is treated as Void. The script sets column M (Mark if Void) = "Void" for that row.
- Protected columns are never overwritten due to a Void.
"""

import logging
import sys
import json
import shutil
from pathlib import Path
from datetime import datetime
from typing import Tuple, List, Dict, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule


class MasterMerger:
    """Merge processed reports into master tracking file with Excel formatting and retention system."""

    def __init__(self, config_path: str = "config.json"):
        self.config_path = Path(config_path)
        self.config = self._load_config()

        # Paths from config.json
        paths = self.config.get("paths", {})
        self.new_reports_dir = Path(paths.get("new_reports_dir", ".")).expanduser().resolve()
        self.master_file = (Path(paths.get("master_dir", ".")) / paths.get(
            "master_file", "Policy4900_Tracking_Master.xlsx")).expanduser().resolve()
        self.overrides_file = paths.get("void_overrides_file", "")

        # Retention file for installed/active cameras
        self.retention_file = Path(self.config.get("paths", {}).get(
            "retention_file",
            str(self.master_file.parent / "Policy4900_Retained.csv")
        )).expanduser().resolve()

        logging.info(f"Reports directory: {self.new_reports_dir}")
        logging.info(f"Master file: {self.master_file}")
        logging.info(f"Retention file: {self.retention_file}")
        if self.overrides_file:
            logging.info(f"Overrides file: {self.overrides_file}")

        # Columns
        columns_cfg = self.config.get("columns", {})
        self.master_columns = columns_cfg.get("master_columns") or [
            "School of Instruction", "FISH Number", "Room", "FISH List",
            "Total Student Count", "# of Students Opt In", "# of Students Opt Out",
            "# of Students No Response", "% Opt In", "% Opt Out", "% No Response",
            "# of ESE Students", "Mark if Void", "Date First Seen", "Change Control",
            "Change Ack", "Change First Seen", "Change Last Seen",
            "Previous Approval Status", "Approval Status", "Date Added to Installation List",
            "Installation Status", "Installation Date", "Activation Status", "Activation Date",
            "Camera Source School", "Camera Source Classroom", "Camera Type", "Notes"
        ]

        # Never overwrite these fields programmatically during merge
        self.protected_columns = [
            "Date First Seen", "Date Added to Installation List",
            "Installation Status", "Installation Date",
            "Activation Status", "Activation Date",
            "Camera Source School", "Camera Source Classroom",
            "Camera Type", "Notes",
            "Change Ack",  # Manual acknowledgment - never overwrite
            "Change Control",  # Preserve Change Control during copy-in
        ]

        self.date_format = self.config.get("output", {}).get("date_format", "%m-%d-%Y")
        self.sheet_name = self.config.get("output", {}).get("master_sheet_name", "Policy 4900 - Classroom Percent")

        # Load manual Void overrides
        self.void_keys: Set[str] = self._load_void_overrides()

    # -------------------- config --------------------

    def _load_config(self) -> dict:
        if not self.config_path.exists():
            logging.warning(f"Config file not found: {self.config_path}. Using defaults.")
            return {"paths": {}, "columns": {}, "output": {}}
        with open(self.config_path, "r", encoding="utf-8") as f:
            return json.load(f)

    # -------------------- backup --------------------

    def backup_master_file(self) -> Path:
        """
        Create a timestamped backup of the master file before processing.
        Returns the path to the backup file, or None if no backup was created.
        """
        if not self.master_file.exists():
            logging.info("No master file to backup (file doesn't exist yet)")
            return None

        # Get backup directory from config, or use default
        backup_dir_str = self.config.get("paths", {}).get("backup_dir", "")
        if not backup_dir_str:
            backup_dir = self.master_file.parent / "Backups"
        else:
            backup_dir = Path(backup_dir_str).expanduser().resolve()

        # Create backup directory if it doesn't exist
        backup_dir.mkdir(parents=True, exist_ok=True)

        # Create timestamp in format: 05-Nov-2025-1430
        timestamp = datetime.now().strftime("%d-%b-%Y-%H%M")

        # Create backup filename: Policy4900_Tracking_Master.backup_05-Nov-2025-1430.xlsx
        backup_name = f"{self.master_file.stem}.backup_{timestamp}{self.master_file.suffix}"
        backup_path = backup_dir / backup_name

        # Copy the file
        try:
            shutil.copy2(self.master_file, backup_path)
            logging.info(f"Master file backed up to: {backup_path}")
            return backup_path
        except Exception as e:
            logging.error(f"Failed to create backup: {e}")
            return None

    # -------------------- retention helpers --------------------

    @staticmethod
    def _is_installed_or_active(row) -> bool:
        """
        Determines if a camera row should be permanently retained.
        Returns True if:
        - Activation Status contains "activat" or "active"
        - Installation Status contains "install" or "existing"
        """
        activation = str(row.get("Activation Status", "")).lower()
        installation = str(row.get("Installation Status", "")).lower()
        return ("activat" in activation) or ("active" in activation) or \
            ("install" in installation) or ("existing" in installation)

    # -------------------- overrides --------------------

    @staticmethod
    def _key(school: str, room: str) -> str:
        return f"{str(school).strip().lower()}||{str(room).strip().lower()}"

    @staticmethod
    def _mk_key_frame(df: pd.DataFrame) -> pd.Series:
        """Create stable merge key: School|Room primary, fallback to School|FISH when Room is blank."""
        school = df["School of Instruction"].astype(str).str.strip().str.lower()
        room = df["Room"].astype(str).str.strip().str.lower()
        fish = df["FISH Number"].astype(str).str.strip().str.lower()
        key = school + "|" + room
        return key.where(room.ne(""), school + "|" + fish)

    def _load_void_overrides(self) -> Set[str]:
        keys: Set[str] = set()
        path = self.overrides_file
        if not path:
            logging.info("No overrides file configured.")
            return keys
        p = Path(path).expanduser()
        if not p.exists():
            logging.warning(f"Overrides file not found: {p}")
            return keys

        df = pd.read_csv(p, dtype=str).fillna("")
        expected_cols = {"School of Instruction", "Room", "Mark if Void", "Approval Status"}
        missing = expected_cols - set(df.columns)
        if missing:
            logging.error(f"Overrides file missing columns: {', '.join(sorted(missing))}")
            return keys

        for _, r in df.iterrows():
            m = str(r.get("Mark if Void", "")).strip().lower()
            q = str(r.get("Approval Status", "")).strip().lower()
            if m == "void" or q == "void":
                keys.add(self._key(r.get("School of Instruction", ""), r.get("Room", "")))

        logging.info(f"Loaded {len(keys)} manual Void override(s).")
        return keys

    # -------------------- file discovery --------------------

    def get_ready_for_merge(self) -> List[Path]:
        pattern = "*_Policy4900_PROCESSED.csv"
        if not self.new_reports_dir.exists():
            logging.error(f"Directory does not exist: {self.new_reports_dir}")
            return []
        all_processed = sorted(self.new_reports_dir.glob(pattern))
        ready = []
        for csv_path in all_processed:
            original_name = csv_path.name.replace("_PROCESSED.csv", ".csv")
            step1 = csv_path.parent / f"{original_name}.step1"
            step2 = csv_path.parent / f"{original_name}.step2"
            if step1.exists() and not step2.exists():
                ready.append(csv_path)
        ready.sort(key=lambda p: self.extract_date_from_filename(p.name))
        return ready

    # -------------------- master I/O with retention --------------------

    def load_master(self) -> pd.DataFrame:
        """
        Load master file with automatic rehydration of retained cameras.
        This ensures installed/active cameras are never lost even on rebuild.
        """
        if not self.master_file.exists():
            logging.info(f"Master file not found. Creating: {self.master_file}")
            df = pd.DataFrame(columns=self.master_columns)
            df["key_sf"] = ""
        else:
            df = pd.read_excel(self.master_file, sheet_name=self.sheet_name, engine="openpyxl")

        # Ensure all columns present
        for col in self.master_columns:
            if col not in df.columns:
                df[col] = ""

        # Ensure key columns are clean strings
        for col in ["School of Instruction", "Room", "FISH Number"]:
            if col not in df.columns:
                df[col] = ""
            df[col] = df[col].astype(str).fillna("").str.strip()

        # Create stable key
        df["key_sf"] = self._mk_key_frame(df)

        for col in ["Previous Approval Status", "Date First Seen", "Change Control",
                    "Change Ack", "Change First Seen", "Change Last Seen"]:
            if col not in df.columns:
                df[col] = ""

        # CRITICAL: Rehydrate retained rows from retention file
        # This prevents data loss if master is rebuilt or cameras disappear temporarily
        if self.retention_file.exists():
            logging.info(f"Loading retained cameras from: {self.retention_file.name}")
            try:
                keep = pd.read_csv(self.retention_file, dtype=str).fillna("")

                # Ensure retention file has all required columns
                for col in self.master_columns:
                    if col not in keep.columns:
                        keep[col] = ""

                # Ensure key columns are clean
                for col in ["School of Instruction", "Room", "FISH Number"]:
                    keep[col] = keep[col].astype(str).fillna("").str.strip()

                keep["key_sf"] = self._mk_key_frame(keep)

                # Merge: Concat and deduplicate (first occurrence wins)
                # This keeps master data if present, adds retention data if missing
                df = pd.concat([df, keep], ignore_index=True)
                df = df.drop_duplicates(subset=["key_sf"], keep="first")

                logging.info(f"Rehydrated {len(keep)} retained camera(s) from retention file")
            except Exception as e:
                logging.warning(f"Could not load retention file: {e}")

        return df

    def _write_master(self, df: pd.DataFrame):
        cols = [c for c in self.master_columns if c in df.columns]
        percent_cols = ["% Opt In", "% Opt Out", "% No Response"]
        df_out = df[cols].copy()
        for col in percent_cols:
            if col in df_out.columns:
                # Only divide by 100 if values are > 1 (whole numbers like 75, not decimals like 0.75)
                s = pd.to_numeric(df_out[col], errors="coerce")
                df_out[col] = s.where(s <= 1, s / 100.0)

        with pd.ExcelWriter(self.master_file, engine="openpyxl", mode="w") as xw:
            df_out.to_excel(xw, index=False, sheet_name=self.sheet_name)
            ws = xw.sheets[self.sheet_name]
            # percentage format
            for col_idx, col_name in enumerate(cols, start=1):
                if col_name in percent_cols:
                    letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, len(df_out) + 2):
                        ws[f"{letter}{row}"].number_format = "0%"

        self._apply_excel_formatting(len(df_out))

    def _update_retention_file(self, df_master: pd.DataFrame):
        """
        Update the retention file with all installed/active cameras.
        This creates a persistent backup that survives master rebuilds.
        """
        # Filter to only installed/active cameras
        retained_mask = df_master.apply(self._is_installed_or_active, axis=1)
        df_keep = df_master.loc[retained_mask, self.master_columns + ["key_sf"]].copy()

        if len(df_keep) == 0:
            logging.info("No installed/active cameras to retain")
            return

        # Merge with existing retention file if it exists
        if self.retention_file.exists():
            try:
                old = pd.read_csv(self.retention_file, dtype=str).fillna("")

                # Ensure all columns present
                for col in self.master_columns:
                    if col not in old.columns:
                        old[col] = ""

                # Ensure key columns clean
                for col in ["School of Instruction", "Room", "FISH Number"]:
                    old[col] = old[col].astype(str).fillna("").str.strip()

                old["key_sf"] = self._mk_key_frame(old)

                # Concat and keep last (most recent data wins)
                merged = pd.concat([old, df_keep], ignore_index=True)
                merged = merged.drop_duplicates(subset=["key_sf"], keep="last")
            except Exception as e:
                logging.warning(f"Could not read existing retention file: {e}")
                merged = df_keep
        else:
            merged = df_keep

        # Write updated retention file
        try:
            merged.to_csv(self.retention_file, index=False, encoding="utf-8")
            logging.info(f"Retention file updated: {len(merged)} camera(s) retained")
        except Exception as e:
            logging.error(f"Failed to write retention file: {e}")

    # -------------------- excel formatting --------------------

    def _apply_excel_formatting(self, data_row_count: int):
        wb = load_workbook(self.master_file)
        ws = wb[self.sheet_name]

        # Map names to letters
        col_letters: Dict[str, str] = {}
        for idx, col_name in enumerate(self.master_columns, start=1):
            # Supports up to 52 columns (A..Z, AA..AZ). Enough for this sheet.
            if idx <= 26:
                col_letters[col_name] = chr(64 + idx)
            else:
                col_letters[col_name] = "A" + chr(64 + idx - 26)

        col_E = col_letters.get("Total Student Count", "E")
        col_F = col_letters.get("# of Students Opt In", "F")
        col_G = col_letters.get("# of Students Opt Out", "G")
        col_H = col_letters.get("# of Students No Response", "H")
        col_M = col_letters.get("Mark if Void", "M")
        col_O = col_letters.get("Change Control", "O")
        col_Q = col_letters.get("Approval Status", "Q")

        # NOTE: Approval Status (column Q) is now calculated in Python, not Excel formulas
        # Python is the single source of truth for this field

        # Conditional formatting
        red_font = Font(color="FF0000")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        green_fill = PatternFill(start_color="66FF99", end_color="66FF99", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Define row range for all conditional formatting rules
        start_row = 2
        end_row = data_row_count + 1
        if end_row < start_row:
            end_row = start_row  # safe clamp for empty/near-empty sheets

        # Column M non-empty -> red text
        rule_m = CellIsRule(operator="notEqual", formula=['""'], font=red_font)
        ws.conditional_formatting.add(f"{col_M}{start_row}:{col_M}{end_row}", rule_m)

        # Column O highlights
        rule_o_red = CellIsRule(operator="equal", formula=['"Missing - Camera Active"'], fill=red_fill)
        rule_o_orange = CellIsRule(operator="equal", formula=['"Approval Lost"'], fill=orange_fill)
        rule_o_green = CellIsRule(operator="equal", formula=['"Approval Gained"'], fill=green_fill)

        # Yellow fill for any "Missing" text
        # NOTE: Formula references top-left cell (col_O + start_row) by design.
        # Excel applies this relatively to each cell in the range, so changing start_row
        # requires updating the formula reference accordingly.
        rule_o_yellow = FormulaRule(formula=[f'ISNUMBER(SEARCH("Missing",{col_O}{start_row}))'], fill=yellow_fill)
        ws.conditional_formatting.add(f"{col_O}{start_row}:{col_O}{end_row}", rule_o_red)
        ws.conditional_formatting.add(f"{col_O}{start_row}:{col_O}{end_row}", rule_o_orange)
        ws.conditional_formatting.add(f"{col_O}{start_row}:{col_O}{end_row}", rule_o_green)
        ws.conditional_formatting.add(f"{col_O}{start_row}:{col_O}{end_row}", rule_o_yellow)

        # Column Q
        rule_q_green = CellIsRule(operator="equal", formula=['"Approved - Camera Authorized"'], fill=green_fill)
        rule_q_denied = CellIsRule(operator="equal", formula=['"Denied - Parent Opt Out"'], font=red_font)
        rule_q_void = CellIsRule(operator="equal", formula=['"Void"'], font=red_font)
        ws.conditional_formatting.add(f"{col_Q}{start_row}:{col_Q}{end_row}", rule_q_green)
        ws.conditional_formatting.add(f"{col_Q}{start_row}:{col_Q}{end_row}", rule_q_denied)
        ws.conditional_formatting.add(f"{col_Q}{start_row}:{col_Q}{end_row}", rule_q_void)

        wb.save(self.master_file)

    # -------------------- approval status calculation --------------------

    @staticmethod
    def _to_int(x):
        """Safe numeric conversion: handles '3.0', 'invalid', None, etc."""
        try:
            # pd.to_numeric handles strings, floats, etc.
            result = pd.to_numeric(x, errors="coerce")
            # If result is NaN (from coerce), use 0
            if pd.isna(result):
                return 0
            return int(result)
        except:
            return 0

    @staticmethod
    def calculate_approval_status(row) -> str:
        """Calculate Approval Status from raw data."""
        # Check for Withdrawn first (hard blocker)
        consent_status = str(row.get("Parent Consent Status", "")).lower()
        withdrawn = "withdraw" in consent_status
        if withdrawn:
            return "Withdrawn"

        # Check for manual Void
        mark_void = str(row.get("Mark if Void", "")).strip().lower()
        if mark_void == "void":
            return "Void"

        # Safe numeric parsing
        total = MasterMerger._to_int(row.get("Total Student Count", 0))
        opt_in = MasterMerger._to_int(row.get("# of Students Opt In", 0))
        opt_out = MasterMerger._to_int(row.get("# of Students Opt Out", 0))
        no_response = MasterMerger._to_int(row.get("# of Students No Response", 0))

        if total == 0:
            return "Not Requested"
        elif opt_out > 0:
            return "Denied - Parent Opt Out"
        elif opt_in == total and total > 0:
            return "Approved - Camera Authorized"
        elif no_response > 0:
            return "Awaiting Responses"
        else:
            return "Not Requested"

    # -------------------- merge core with retention --------------------

    def merge_report_into_master(
            self, df_master: pd.DataFrame, df_new: pd.DataFrame, report_date: str
    ) -> Tuple[pd.DataFrame, dict]:
        # Ensure keys exist with stable School|Room primary, School|FISH fallback
        for d in (df_master, df_new):
            for col in ["School of Instruction", "Room", "FISH Number"]:
                if col not in d.columns:
                    d[col] = ""
                d[col] = d[col].astype(str).fillna("").str.strip()
            d["key_sf"] = self._mk_key_frame(d)

        master_idx_by_key = {k: i for i, k in enumerate(df_master["key_sf"])}
        new_keys = set(df_new["key_sf"])

        stats = {
            "added": 0, "approval_gained": 0, "approval_lost": 0,
            "status_changed": 0, "no_change": 0,
            "missing_camera_active": 0, "missing_camera_installed": 0,
            "missing_was_approved": 0, "missing_no_camera": 0,
            "manual_void_applied": 0,
            # Lists to track specific classrooms for reporting
            "approval_gained_list": [],
            "approval_lost_list": [],
            "missing_was_approved_list": [],
            "missing_camera_active_list": [],
            "missing_camera_installed_list": []
        }

        # Step 1: apply incoming rows
        for _, row in df_new.iterrows():
            k = row["key_sf"]
            if k in master_idx_by_key:
                i = master_idx_by_key[k]

                # CRITICAL: Calculate CURRENT Approval Status from existing master data BEFORE updating
                old_approval = self.calculate_approval_status(df_master.iloc[i])
                df_master.at[i, "Previous Approval Status"] = old_approval

                # Update non-protected columns from report
                for col in self.master_columns:
                    if col in row.index and col not in self.protected_columns:
                        val = row[col]
                        if pd.notna(val) and str(val).strip() != "":
                            df_master.at[i, col] = val

                # Calculate NEW Approval Status from updated data
                new_approval = self.calculate_approval_status(df_master.iloc[i])
                df_master.at[i, "Approval Status"] = new_approval

                # If approved and first time, set Date Added to Installation List
                if new_approval == "Approved - Camera Authorized":
                    existing = df_master.at[i, "Date Added to Installation List"]
                    if pd.isna(existing) or str(existing).strip() == "":
                        df_master.at[i, "Date Added to Installation List"] = report_date

                # Change Control with one-time ACK behavior
                ack = str(df_master.at[i, "Change Ack"] or "").strip()

                transition_gained = (new_approval == "Approved - Camera Authorized"
                                     and old_approval != "Approved - Camera Authorized")
                transition_lost = (old_approval == "Approved - Camera Authorized"
                                   and new_approval != "Approved - Camera Authorized")

                if transition_gained or transition_lost:
                    # Always flag new transitions
                    label = "Approval Gained" if transition_gained else "Approval Lost"
                    df_master.at[i, "Change Control"] = label
                    if not str(df_master.at[i, "Change First Seen"] or "").strip():
                        df_master.at[i, "Change First Seen"] = report_date
                    df_master.at[i, "Change Last Seen"] = report_date
                    # Clear any old ACK; it was for a prior change
                    if ack:
                        df_master.at[i, "Change Ack"] = ""

                    # Track which classroom changed with report date
                    school = str(df_master.at[i, "School of Instruction"] or "").strip()
                    room = str(df_master.at[i, "Room"] or "").strip()
                    classroom_info = f"{school} - Room {room} (Report: {report_date})"

                    if transition_gained:
                        stats["approval_gained"] += 1
                        stats["approval_gained_list"].append(classroom_info)
                    else:
                        stats["approval_lost"] += 1
                        stats["approval_lost_list"].append(classroom_info)
                else:
                    # No new transition this run
                    if ack:
                        # Consume ACK once, clear flag, then clear ACK
                        df_master.at[i, "Change Control"] = "No Change"
                        df_master.at[i, "Change Ack"] = ""
                        stats["no_change"] += 1
                    else:
                        cur = str(df_master.at[i, "Change Control"] or "").strip()
                        if cur in {"Approval Gained", "Approval Lost"}:
                            # Keep sticky alert visible until acknowledged
                            df_master.at[i, "Change Last Seen"] = report_date
                        elif old_approval != new_approval:
                            df_master.at[i, "Change Control"] = "Status Changed"
                            stats["status_changed"] += 1
                        else:
                            df_master.at[i, "Change Control"] = "No Change"
                            stats["no_change"] += 1

            else:
                # New classroom
                new_row = {c: row.get(c, "") for c in self.master_columns}
                new_row["key_sf"] = k
                new_row["Change Control"] = "Added"
                new_row["Previous Approval Status"] = ""
                new_row["Date First Seen"] = report_date

                # Calculate Approval Status for new classroom
                new_row["Approval Status"] = self.calculate_approval_status(row)

                if new_row.get("Approval Status") == "Approved - Camera Authorized":
                    new_row["Date Added to Installation List"] = report_date

                df_master = pd.concat([df_master, pd.DataFrame([new_row])], ignore_index=True)
                master_idx_by_key[k] = len(df_master) - 1
                stats["added"] += 1

        # Step 2: apply manual Void overrides to master
        if self.void_keys:
            for i, r in df_master.iterrows():
                key = self._key(r.get("School of Instruction", ""), r.get("Room", ""))
                if key in self.void_keys:
                    if df_master.at[i, "Mark if Void"] != "Void":
                        df_master.at[i, "Mark if Void"] = "Void"
                        # Recalculate Approval Status after marking as Void
                        df_master.at[i, "Approval Status"] = self.calculate_approval_status(df_master.iloc[i])
                        stats["manual_void_applied"] += 1
                    # Do not overwrite protected columns

        # Step 3: mark rows missing from the new report
        # CRITICAL ENHANCEMENT: Never drop installed/active cameras
        # NOTE: Missing status is PERSISTENT - shows what's currently missing,
        # not a one-time change event. It stays until classroom returns to reports.
        missing_mask = ~df_master["key_sf"].isin(new_keys)
        for i in df_master[missing_mask].index:
            # More robust status checking with null-safety
            activation_status = str(df_master.at[i, "Activation Status"] or "").strip().lower()
            installation_status = str(df_master.at[i, "Installation Status"] or "").strip().lower()
            approval_status = str(df_master.at[i, "Approval Status"] or "").strip().lower()
            prev_approval = str(df_master.at[i, "Previous Approval Status"] or "").strip().lower()

            # Enhanced classification with flexible substring matching
            # CRITICAL: Check for both "activat" and "active" to catch both "Activated" and "Active"
            school = str(df_master.at[i, "School of Instruction"] or "").strip()
            room = str(df_master.at[i, "Room"] or "").strip()
            classroom_info = f"{school} - Room {room} (Report: {report_date})"

            if ("activat" in activation_status) or ("active" in activation_status):
                label = "Missing - Camera Active"
                stats["missing_camera_active"] += 1
                stats["missing_camera_active_list"].append(classroom_info)
            elif ("install" in installation_status) or ("existing" in installation_status):
                label = "Missing - Camera Installed"
                stats["missing_camera_installed"] += 1
                stats["missing_camera_installed_list"].append(classroom_info)
            elif ("approved" in approval_status) or ("approved" in prev_approval):
                label = "Missing - Was Approved"
                stats["missing_was_approved"] += 1
                stats["missing_was_approved_list"].append(classroom_info)
            else:
                label = "Missing - No Camera"
                stats["missing_no_camera"] += 1

            # Persistent state: always show the missing label
            df_master.at[i, "Change Control"] = label

            # Track when this missing state was first seen
            if not str(df_master.at[i, "Change First Seen"] or "").strip():
                df_master.at[i, "Change First Seen"] = report_date
            df_master.at[i, "Change Last Seen"] = report_date

            # ACK not used for missing status (it's persistent, not a change event)
            if str(df_master.at[i, "Change Ack"] or "").strip():
                df_master.at[i, "Change Ack"] = ""

        return df_master, stats

    # -------------------- markers --------------------

    def create_step2_marker(self, processed_path: Path, report_date: str, stats: dict):
        original_name = processed_path.name.replace("_PROCESSED.csv", ".csv")
        marker_path = processed_path.parent / f"{original_name}.step2"
        metadata = {
            "processed_file": processed_path.name,
            "merged_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "report_date": report_date,
            "statistics": stats,
        }
        with open(marker_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2)
        logging.info(f"Created marker: {marker_path.name}")
        return marker_path

    # -------------------- helpers --------------------

    @staticmethod
    def extract_date_from_filename(name: str) -> datetime:
        stem = Path(name).stem.replace("_PROCESSED", "")
        return datetime.strptime(stem[:10], "%m-%d-%Y")


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    print("=" * 70)
    print("TASK 2: Merge Processed Reports into Master Tracking")
    print("(Enhanced with Retention System for Installed/Active Cameras)")
    print("=" * 70)
    print()

    merger = MasterMerger()
    print(f"[SEARCH] Master file: {merger.master_file.name}")
    print(f"[SEARCH] Retention file: {merger.retention_file.name}")
    print(f"[DIR] Reports directory: {merger.new_reports_dir}")
    if merger.overrides_file:
        print(f"[VOID] Void overrides: {merger.overrides_file}")
    print()

    ready_files = merger.get_ready_for_merge()
    if not ready_files:
        print("[OK] No new reports to process")
        print("  (All reports have already been merged)")
        print()
        print("=" * 70)
        return

    # CRITICAL: Backup master file before any processing
    print("[BACKUP] Creating backup of master file...")
    backup_path = merger.backup_master_file()
    if backup_path:
        print(f"[OK] Backup created: {backup_path.name}")
    else:
        print("[INFO] No backup needed (master file doesn't exist yet)")
    print()

    try:
        df_master = merger.load_master()
        initial_rows = len(df_master)
        print(f"[STATS] Current master file: {initial_rows} classroom(s)")
    except Exception as e:
        logging.error(f"Failed to load master: {e}")
        sys.exit(1)

    totals = {
        "added": 0, "approval_gained": 0, "approval_lost": 0, "status_changed": 0, "no_change": 0,
        "missing_camera_active": 0, "missing_camera_installed": 0, "missing_was_approved": 0,
        "missing_no_camera": 0, "manual_void_applied": 0,
        "approval_gained_list": [],
        "approval_lost_list": [],
        "missing_was_approved_list": [],
        "missing_camera_active_list": [],
        "missing_camera_installed_list": []
    }
    merged = 0
    failed = 0

    # Process each report individually
    print(f"\nProcessing {len(ready_files)} report(s) in chronological order...")
    print()

    for report_num, processed_path in enumerate(ready_files, 1):
        try:
            report_date = merger.extract_date_from_filename(processed_path.name).strftime(merger.date_format)

            print(f"{'=' * 70}")
            print(f"Report {report_num}/{len(ready_files)}: {processed_path.name}")
            print(f"Date: {report_date}")
            print(f"{'=' * 70}")

            # Load the report
            # Handle potential encoding issues in CSVs with retry logic
            try:
                df_new = pd.read_csv(processed_path, encoding="utf-8")
            except UnicodeDecodeError:
                logging.warning(f"UTF-8 decode failed for {processed_path.name}, retrying with error handling")
                df_new = pd.read_csv(processed_path, encoding="utf-8", encoding_errors="ignore")
            print(f"Classrooms in report: {len(df_new)}")

            # Merge into master
            df_master, stats = merger.merge_report_into_master(df_master, df_new, report_date)

            # Update cumulative totals
            for k in totals:
                if isinstance(totals[k], list):
                    # For lists, extend rather than add
                    totals[k].extend(stats.get(k, []))
                else:
                    # For numbers, add as before
                    totals[k] += stats.get(k, 0)

            # Show this report's changes
            print(f"\nChanges detected:")
            if stats['added'] > 0:
                print(f"  [OK] Added: {stats['added']}")
            if stats['approval_gained'] > 0:
                print(f"  [OK] Approval Gained: {stats['approval_gained']}")
                # Show which classrooms gained approval
                for classroom in stats['approval_gained_list']:
                    print(f"    - {classroom}")
            if stats['approval_lost'] > 0:
                print(f"  [WARNING]  Approval Lost: {stats['approval_lost']}")
                # Show which classrooms lost approval
                for classroom in stats['approval_lost_list']:
                    print(f"    - {classroom}")
            if stats['status_changed'] > 0:
                print(f"  * Status Changed: {stats['status_changed']}")
            if stats['no_change'] > 0:
                print(f"  * No Change: {stats['no_change']}")
            if stats['manual_void_applied'] > 0:
                print(f"  * Manual Void Applied: {stats['manual_void_applied']}")

            missing_this_report = (
                    stats['missing_camera_active'] + stats['missing_camera_installed'] +
                    stats['missing_was_approved'] + stats['missing_no_camera']
            )
            if missing_this_report > 0:
                print(f"\n[WARNING]  Classrooms missing from this report: {missing_this_report}")
                if stats['missing_camera_active'] > 0:
                    print(f"    [CRITICAL] CRITICAL - Camera Active but missing: {stats['missing_camera_active']}")
                    for classroom in stats['missing_camera_active_list']:
                        print(f"      - {classroom}")
                if stats['missing_camera_installed'] > 0:
                    print(f"    * Camera Installed but missing: {stats['missing_camera_installed']}")
                    for classroom in stats['missing_camera_installed_list']:
                        print(f"      - {classroom}")
                if stats['missing_was_approved'] > 0:
                    print(f"    * Was Approved but missing: {stats['missing_was_approved']}")
                    for classroom in stats['missing_was_approved_list']:
                        print(f"      - {classroom}")
                if stats['missing_no_camera'] > 0:
                    print(f"    * No Camera, just missing: {stats['missing_no_camera']}")

            # CRITICAL: Write master file AFTER EACH REPORT
            # This preserves Change Control values specific to this report's date
            # Without this, the next report would overwrite these changes!
            print(f"\nSaving master file...")
            merger._write_master(df_master)
            print(f"[OK] Master file updated for {report_date}")

            # CRITICAL: Update retention file with all installed/active cameras
            # This creates a persistent backup that survives master rebuilds
            print(f"Updating retention file...")
            merger._update_retention_file(df_master)
            print(f"[OK] Retention file updated")

            # Create step2 marker (prevents reprocessing)
            merger.create_step2_marker(processed_path, report_date, stats)
            print(f"[OK] Marker created")

            # Reload master for next report
            # This ensures we start with the saved state with proper formatting
            if report_num < len(ready_files):
                df_master = merger.load_master()

            merged += 1
            print()

        except Exception as e:
            print(f"\n[ERROR] ERROR processing {processed_path.name}")
            logging.exception(e)
            failed += 1
            print()

    # Final summary
    print("=" * 70)
    print("TASK 2 SUMMARY")
    print("=" * 70)

    if merged > 0:
        try:
            final_rows = len(df_master)

            print(f"\n[OK] Successfully processed {merged} report(s)")
            if failed > 0:
                print(f"[ERROR] Failed: {failed} report(s)")

            print(f"\nMaster file status:")
            print(f"  Initial rows: {initial_rows}")
            print(f"  Final rows: {final_rows}")
            print(f"  Net change: +{final_rows - initial_rows}")

            print(f"\n[STATS] Cumulative Statistics (across all {merged} reports):")
            print(f"  * Added (new classrooms): {totals['added']}")
            if totals['approval_gained'] > 0:
                print(f"  * Approval Gained: {totals['approval_gained']}")
                for classroom in totals['approval_gained_list']:
                    print(f"    - {classroom}")
            else:
                print(f"  * Approval Gained: {totals['approval_gained']}")
            if totals['approval_lost'] > 0:
                print(f"  * Approval Lost: {totals['approval_lost']}")
                for classroom in totals['approval_lost_list']:
                    print(f"    - {classroom}")
            else:
                print(f"  * Approval Lost: {totals['approval_lost']}")
            print(f"  * Status Changed: {totals['status_changed']}")
            print(f"  * No Change: {totals['no_change']}")
            if totals['manual_void_applied'] > 0:
                print(f"  * Manual Void Applied: {totals['manual_void_applied']}")

            missing_total = (
                    totals['missing_camera_active'] + totals['missing_camera_installed'] +
                    totals['missing_was_approved'] + totals['missing_no_camera']
            )
            if missing_total > 0:
                print(f"\n[WARNING]  Missing Classroom Alerts (cumulative): {missing_total}")
                if totals['missing_camera_active'] > 0:
                    print(f"  [CRITICAL] CRITICAL - Camera Active but missing: {totals['missing_camera_active']}")
                    for classroom in totals['missing_camera_active_list']:
                        print(f"    - {classroom}")
                if totals['missing_camera_installed'] > 0:
                    print(f"  * Camera Installed but missing: {totals['missing_camera_installed']}")
                    for classroom in totals['missing_camera_installed_list']:
                        print(f"    - {classroom}")
                if totals['missing_was_approved'] > 0:
                    print(f"  * Was Approved but missing: {totals['missing_was_approved']}")
                    for classroom in totals['missing_was_approved_list']:
                        print(f"    - {classroom}")
                if totals['missing_no_camera'] > 0:
                    print(f"  * No Camera, just missing: {totals['missing_no_camera']}")

            print(f"\n[OK] Master file: {merger.master_file}")
            print(f"[OK] Retention file: {merger.retention_file}")
            print(f"[OK] Includes: Python-calculated Approval Status, Conditional Formatting, Manual Void Overrides")
            print(f"\n[TIP] Next step: Run your VBA macro for borders, freeze panes, and column widths")
            print("=" * 70)

        except Exception as e:
            logging.error(f"Failed to generate summary: {e}")
            sys.exit(1)
    else:
        print(f"\n[WARNING]  No reports were successfully merged")
        if failed > 0:
            print(f"[ERROR] {failed} report(s) failed")
        print("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except Exception as ex:
        logging.exception(ex)
        sys.exit(1)